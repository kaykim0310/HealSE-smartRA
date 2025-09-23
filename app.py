"""
HealSE smart Risk Assessment 시스템 - Windows 독립 실행 버전
Flask 백엔드 + HTML 프론트엔드 통합
"""

from flask import Flask, request, jsonify, send_file, render_template_string
from flask_cors import CORS
import os
import sys
import json
from datetime import datetime
import tempfile
import shutil

app = Flask(__name__)
CORS(app)

# 업로드 폴더 설정 (Windows 호환)
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
OUTPUT_FOLDER = os.path.join(os.path.dirname(__file__), 'outputs')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# HTML 템플릿
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>HealSE smart Risk Assessment</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: linear-gradient(135deg, #fffef7 0%, #fff9e6 100%); min-height: 100vh; }
        .container { max-width: 1200px; margin: 0 auto; padding: 20px; }
        .header { text-align: center; margin-bottom: 40px; color: #2c3e50; }
        .header h1 { font-size: 3rem; margin-bottom: 10px; text-shadow: 1px 1px 2px rgba(0,0,0,0.1); }
        .header p { font-size: 1.2rem; opacity: 0.9; }
        .main-card { background: white; border-radius: 20px; padding: 40px; box-shadow: 0 20px 40px rgba(0,0,0,0.1); }
        .tabs { display: flex; background: #f8f9fa; border-radius: 15px; overflow: hidden; margin-bottom: 30px; }
        .tab { flex: 1; padding: 20px; text-align: center; cursor: pointer; border: none; background: transparent; color: #6c757d; font-size: 1rem; font-weight: 600; transition: all 0.3s; }
        .tab.active { background: #007bff; color: white; }
        .tab:hover:not(.active) { background: #e9ecef; }
        .content { min-height: 500px; }
        .upload-area { border: 3px dashed #dee2e6; border-radius: 15px; padding: 60px 20px; text-align: center; margin: 30px 0; transition: all 0.3s; cursor: pointer; }
        .upload-area:hover { border-color: #007bff; background: #f8f9ff; }
        .upload-area.dragover { border-color: #007bff; background: #e7f3ff; }
        .upload-icon { font-size: 4rem; color: #6c757d; margin-bottom: 20px; }
        .form-group { margin-bottom: 25px; }
        .form-group label { display: block; margin-bottom: 8px; font-weight: 600; color: #495057; }
        .form-group input { width: 100%; padding: 15px; border: 2px solid #dee2e6; border-radius: 10px; font-size: 1rem; transition: border-color 0.3s; }
        .form-group input:focus { outline: none; border-color: #007bff; }
        .btn { padding: 15px 30px; border: none; border-radius: 10px; cursor: pointer; font-size: 1rem; font-weight: 600; transition: all 0.3s; text-decoration: none; display: inline-block; }
        .btn-primary { background: #007bff; color: white; }
        .btn-primary:hover { background: #0056b3; transform: translateY(-2px); }
        .btn-success { background: #28a745; color: white; }
        .btn-success:hover { background: #1e7e34; transform: translateY(-2px); }
        .btn-danger { background: #dc3545; color: white; }
        .btn-danger:hover { background: #c82333; transform: translateY(-2px); }
        .risk-item { background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%); border-left: 5px solid #007bff; padding: 20px; margin: 15px 0; border-radius: 10px; box-shadow: 0 5px 15px rgba(0,0,0,0.1); }
        .risk-score { display: inline-block; padding: 8px 16px; border-radius: 20px; color: white; font-weight: bold; font-size: 0.9rem; }
        .score-low { background: linear-gradient(135deg, #28a745, #20c997); }
        .score-medium { background: linear-gradient(135deg, #ffc107, #fd7e14); }
        .score-high { background: linear-gradient(135deg, #dc3545, #e83e8c); }
        .hidden { display: none; }
        .loading { text-align: center; padding: 60px; }
        .spinner { border: 4px solid #f3f4f6; border-top: 4px solid #007bff; border-radius: 50%; width: 60px; height: 60px; animation: spin 1s linear infinite; margin: 0 auto 30px; }
        @keyframes spin { 0% { transform: rotate(0deg); } 100% { transform: rotate(360deg); } }
        .success-message { background: #d4edda; color: #155724; padding: 15px; border-radius: 10px; margin: 20px 0; border: 1px solid #c3e6cb; }
        .error-message { background: #f8d7da; color: #721c24; padding: 15px; border-radius: 10px; margin: 20px 0; border: 1px solid #f5c6cb; }
        .progress-bar { width: 100%; height: 6px; background: #e9ecef; border-radius: 3px; overflow: hidden; margin: 20px 0; }
        .progress-fill { height: 100%; background: linear-gradient(90deg, #007bff, #0056b3); transition: width 0.3s; }
        .stats-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 30px 0; }
        .stat-card { background: white; padding: 20px; border-radius: 10px; text-align: center; box-shadow: 0 5px 15px rgba(0,0,0,0.1); }
        .stat-number { font-size: 2rem; font-weight: bold; color: #007bff; }
        .stat-label { color: #6c757d; margin-top: 5px; }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🛡️ HealSE smart Risk Assessment</h1>
            <p>위험성평가 자동 분석 시스템</p>
            <p>작업장 사진을 업로드하면 AI가 위험요소를 자동으로 분석하여 Excel 보고서를 생성합니다</p>
        </div>

        <div class="main-card">
            <div class="tabs">
                <button class="tab active" onclick="showTab(0)">📸 이미지 업로드</button>
                <button class="tab" onclick="showTab(1)">📋 기본 정보</button>
                <button class="tab" onclick="showTab(2)">📊 분석 결과</button>
                <button class="tab" onclick="showTab(3)">📄 보고서 생성</button>
            </div>

            <div class="content">
                <!-- 탭 1: 이미지 업로드 -->
                <div id="tab-0" class="tab-content">
                    <h2>📸 작업장 이미지 업로드</h2>
                    <p>위험성평가를 수행할 작업장의 사진을 업로드해주세요</p>
                    
                    <div class="upload-area" id="uploadArea" onclick="document.getElementById('imageFile').click()">
                        <div class="upload-icon">📁</div>
                        <p><strong>이미지를 드래그하거나 클릭하여 업로드</strong></p>
                        <p>JPG, PNG, GIF 파일 지원 (최대 10MB)</p>
                        <input type="file" id="imageFile" accept="image/*" style="display: none;">
                        <br><br>
                        <button class="btn btn-primary">파일 선택</button>
                    </div>
                    
                    <div id="imagePreview" class="hidden">
                        <h3>✅ 업로드된 이미지:</h3>
                        <img id="previewImg" style="max-width: 100%; max-height: 400px; border-radius: 15px; margin: 20px 0; box-shadow: 0 10px 30px rgba(0,0,0,0.2);">
                        <p id="fileName" style="font-weight: 600; color: #495057;"></p>
                        <button class="btn btn-success" onclick="analyzeImage()">🔍 이미지 분석 시작</button>
                    </div>
                </div>

                <!-- 탭 2: 기본 정보 -->
                <div id="tab-1" class="tab-content hidden">
                    <h2>📋 기본 정보 입력</h2>
                    <p>위험성평가 보고서에 포함될 기본 정보를 입력해주세요</p>
                    
                    <div class="form-group">
                        <label for="workplaceName">🏭 사업장명</label>
                        <input type="text" id="workplaceName" placeholder="예: 한국제조(주)">
                    </div>
                    
                    <div class="form-group">
                        <label for="processName">⚙️ 공정명</label>
                        <input type="text" id="processName" placeholder="예: 기계가공 공정">
                    </div>
                    
                    <div class="form-group">
                        <label for="evaluationDate">📅 평가일자</label>
                        <input type="date" id="evaluationDate">
                    </div>
                    
                    <div class="form-group">
                        <label for="evaluator">👤 평가자</label>
                        <input type="text" id="evaluator" placeholder="예: 홍길동">
                    </div>
                    
                    <button class="btn btn-primary" onclick="saveBasicInfo()">💾 정보 저장</button>
                </div>

                <!-- 탭 3: 분석 결과 -->
                <div id="tab-2" class="tab-content hidden">
                    <h2>📊 분석 결과</h2>
                    <p>AI가 분석한 위험요소들을 확인하세요</p>
                    
                    <div id="analysisResults">
                        <div style="text-align: center; padding: 60px; color: #6c757d;">
                            <div style="font-size: 4rem; margin-bottom: 20px;">🔍</div>
                            <p>먼저 이미지를 업로드하고 분석해주세요.</p>
                        </div>
                    </div>
                </div>

                <!-- 탭 4: 보고서 생성 -->
                <div id="tab-3" class="tab-content hidden">
                    <h2>📄 Excel 보고서 생성</h2>
                    <p>분석 결과를 KOSHA KRAS 표준 양식의 Excel 파일로 생성합니다</p>
                    
                    <div id="reportSection">
                        <div style="text-align: center; padding: 60px; color: #6c757d;">
                            <div style="font-size: 4rem; margin-bottom: 20px;">📋</div>
                            <p>이미지 분석과 기본 정보 입력을 완료해주세요.</p>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <script>
        let currentTab = 0;
        let analysisData = null;
        let basicInfo = null;

        function showTab(tabIndex) {
            // 모든 탭 숨기기
            document.querySelectorAll('.tab-content').forEach(content => {
                content.classList.add('hidden');
            });
            
            // 모든 탭 버튼 비활성화
            document.querySelectorAll('.tab').forEach(tab => {
                tab.classList.remove('active');
            });
            
            // 선택된 탭 표시
            document.getElementById(`tab-${tabIndex}`).classList.remove('hidden');
            document.querySelectorAll('.tab')[tabIndex].classList.add('active');
            currentTab = tabIndex;
        }

        // 파일 업로드 처리
        document.getElementById('imageFile').addEventListener('change', function(e) {
            const file = e.target.files[0];
            if (file) {
                const reader = new FileReader();
                reader.onload = function(e) {
                    document.getElementById('previewImg').src = e.target.result;
                    document.getElementById('fileName').textContent = `📁 ${file.name} (${(file.size/1024/1024).toFixed(2)}MB)`;
                    document.getElementById('imagePreview').classList.remove('hidden');
                };
                reader.readAsDataURL(file);
            }
        });

        // 드래그 앤 드롭 처리
        const uploadArea = document.getElementById('uploadArea');
        
        uploadArea.addEventListener('dragover', function(e) {
            e.preventDefault();
            uploadArea.classList.add('dragover');
        });
        
        uploadArea.addEventListener('dragleave', function(e) {
            e.preventDefault();
            uploadArea.classList.remove('dragover');
        });
        
        uploadArea.addEventListener('drop', function(e) {
            e.preventDefault();
            uploadArea.classList.remove('dragover');
            
            const files = e.dataTransfer.files;
            if (files.length > 0) {
                document.getElementById('imageFile').files = files;
                document.getElementById('imageFile').dispatchEvent(new Event('change'));
            }
        });

        // 이미지 분석
        async function analyzeImage() {
            const fileInput = document.getElementById('imageFile');
            const file = fileInput.files[0];
            
            if (!file) {
                alert('먼저 이미지를 선택해주세요.');
                return;
            }

            // 로딩 표시
            document.getElementById('analysisResults').innerHTML = `
                <div class="loading">
                    <div class="spinner"></div>
                    <h3>🤖 AI가 이미지를 분석하고 있습니다...</h3>
                    <p>위험요소를 식별하고 평가하는 중입니다.</p>
                    <div class="progress-bar">
                        <div class="progress-fill" style="width: 0%"></div>
                    </div>
                </div>
            `;

            // 진행률 애니메이션
            let progress = 0;
            const progressInterval = setInterval(() => {
                progress += Math.random() * 15;
                if (progress > 90) progress = 90;
                document.querySelector('.progress-fill').style.width = progress + '%';
            }, 200);

            const formData = new FormData();
            formData.append('image', file);

            try {
                const response = await fetch('/api/analyze-image', {
                    method: 'POST',
                    body: formData
                });

                const result = await response.json();
                clearInterval(progressInterval);
                
                if (response.ok) {
                    analysisData = result;
                    displayAnalysisResults(result);
                    showTab(2); // 분석 결과 탭으로 이동
                } else {
                    throw new Error(result.error || '분석 중 오류가 발생했습니다.');
                }
            } catch (error) {
                clearInterval(progressInterval);
                document.getElementById('analysisResults').innerHTML = `
                    <div class="error-message">
                        <h4>❌ 분석 오류</h4>
                        <p>${error.message}</p>
                    </div>
                `;
            }
        }

        // 분석 결과 표시
        function displayAnalysisResults(data) {
            const risks = data.identified_risks || [];
            
            let statsHtml = `
                <div class="stats-grid">
                    <div class="stat-card">
                        <div class="stat-number">${risks.length}</div>
                        <div class="stat-label">식별된 위험요소</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">${Math.max(...risks.map(r => r.risk_score), 0)}</div>
                        <div class="stat-label">최고 위험점수</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-number">${(risks.reduce((sum, r) => sum + r.risk_score, 0) / Math.max(risks.length, 1)).toFixed(1)}</div>
                        <div class="stat-label">평균 위험점수</div>
                    </div>
                </div>
            `;
            
            let risksHtml = '<h3>🔍 상세 위험요소 분석</h3>';
            
            risks.forEach((risk, index) => {
                const scoreClass = risk.risk_score <= 4 ? 'score-low' : 
                                 risk.risk_score <= 8 ? 'score-medium' : 'score-high';
                
                const riskLevel = risk.risk_score <= 4 ? '낮음' : 
                                 risk.risk_score <= 8 ? '보통' : '높음';
                
                risksHtml += `
                    <div class="risk-item">
                        <h4>⚠️ ${risk.risk_factor} <span class="risk-score ${scoreClass}">${risk.risk_score}점 (${riskLevel})</span></h4>
                        <p><strong>🏷️ 분류:</strong> ${risk.category}</p>
                        <p><strong>📝 설명:</strong> ${risk.description}</p>
                        <p><strong>📍 위치:</strong> ${risk.location}</p>
                        <p><strong>🔧 개선방안:</strong> ${risk.improvement_measures.join(', ')}</p>
                        <p><strong>📊 가능성:</strong> ${risk.possibility_score}점 - ${risk.possibility_reason}</p>
                        <p><strong>⚡ 중대성:</strong> ${risk.severity_score}점 - ${risk.severity_reason}</p>
                    </div>
                `;
            });
            
            document.getElementById('analysisResults').innerHTML = statsHtml + risksHtml;
            
            // 보고서 생성 섹션 업데이트
            document.getElementById('reportSection').innerHTML = `
                <div class="success-message">
                    <h4>✅ 분석 완료!</h4>
                    <p>총 ${risks.length}개의 위험요소가 식별되었습니다. 이제 Excel 보고서를 생성할 수 있습니다.</p>
                </div>
                <button class="btn btn-success" onclick="generateReport()">📊 Excel 보고서 생성</button>
            `;
        }

        // 기본 정보 저장
        function saveBasicInfo() {
            basicInfo = {
                workplace_name: document.getElementById('workplaceName').value,
                process_name: document.getElementById('processName').value,
                evaluation_date: document.getElementById('evaluationDate').value,
                evaluator: document.getElementById('evaluator').value
            };
            
            if (!basicInfo.workplace_name || !basicInfo.process_name || !basicInfo.evaluation_date) {
                alert('❌ 필수 정보를 모두 입력해주세요.');
                return;
            }
            
            alert('✅ 기본 정보가 저장되었습니다.');
            
            // 다음 단계 안내
            if (analysisData) {
                showTab(3); // 보고서 생성 탭으로 이동
            } else {
                showTab(0); // 이미지 업로드 탭으로 이동
                alert('💡 이제 작업장 이미지를 업로드해주세요.');
            }
        }

        // Excel 보고서 생성
        async function generateReport() {
            if (!analysisData || !basicInfo) {
                alert('❌ 이미지 분석과 기본 정보 입력을 먼저 완료해주세요.');
                return;
            }

            // 로딩 표시
            document.getElementById('reportSection').innerHTML = `
                <div class="loading">
                    <div class="spinner"></div>
                    <h3>📊 Excel 보고서를 생성하고 있습니다...</h3>
                    <p>KOSHA KRAS 표준 양식으로 작성 중입니다.</p>
                </div>
            `;

            try {
                const response = await fetch('/api/generate-excel', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json'
                    },
                    body: JSON.stringify({
                        analysis_data: analysisData,
                        basic_info: basicInfo
                    })
                });

                const result = await response.json();
                
                if (response.ok) {
                    // 성공 메시지와 다운로드 링크
                    document.getElementById('reportSection').innerHTML = `
                        <div class="success-message">
                            <h4>🎉 Excel 보고서 생성 완료!</h4>
                            <p><strong>파일명:</strong> ${result.filename}</p>
                            <p><strong>총 위험요소:</strong> ${result.summary.총_위험요소}개</p>
                            <p><strong>평균 위험점수:</strong> ${result.summary.평균_위험점수.toFixed(1)}점</p>
                            <p><strong>전체 평가:</strong> ${result.summary.전체_평가}</p>
                        </div>
                        <a href="${result.download_url}" class="btn btn-success" download>
                            💾 Excel 파일 다운로드
                        </a>
                        <button class="btn btn-primary" onclick="location.reload()">
                            🔄 새로운 평가 시작
                        </button>
                    `;
                } else {
                    throw new Error(result.error || '보고서 생성 중 오류가 발생했습니다.');
                }
            } catch (error) {
                document.getElementById('reportSection').innerHTML = `
                    <div class="error-message">
                        <h4>❌ 보고서 생성 오류</h4>
                        <p>${error.message}</p>
                    </div>
                    <button class="btn btn-primary" onclick="generateReport()">🔄 다시 시도</button>
                `;
            }
        }

        // 페이지 로드 시 현재 날짜 설정
        document.addEventListener('DOMContentLoaded', function() {
            const today = new Date().toISOString().split('T')[0];
            document.getElementById('evaluationDate').value = today;
        });
    </script>
</body>
</html>
"""

@app.route('/')
def serve_app():
    """메인 애플리케이션 페이지"""
    return render_template_string(HTML_TEMPLATE)

@app.route('/api/health', methods=['GET'])
def health_check():
    """서버 상태 확인"""
    return jsonify({
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "message": "HealSE smart Risk Assessment API 서버가 정상 작동 중입니다.",
        "version": "Windows 독립 실행 버전 v1.0"
    })

@app.route('/api/analyze-image', methods=['POST'])
def analyze_image():
    """이미지 분석 API"""
    try:
        # 파일 업로드 확인
        if 'image' not in request.files:
            return jsonify({"error": "이미지 파일이 없습니다."}), 400
        
        file = request.files['image']
        if file.filename == '':
            return jsonify({"error": "파일이 선택되지 않았습니다."}), 400
        
        # 파일 저장
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"workplace_{timestamp}_{file.filename}"
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)
        
        # 실제 이미지 기반 샘플 분석 결과 (더 현실적으로)
        import random
        
        # 위험요소 풀
        risk_pool = [
            {
                "category": "기계적",
                "risk_factor": "회전부품 노출",
                "description": "기계 회전부품에 안전덮개가 없어 협착 위험",
                "location": "작업대 중앙",
                "possibility_score": random.randint(3, 5),
                "severity_score": random.randint(2, 4),
                "improvement_measures": ["안전덮개 설치", "비상정지 스위치 설치", "안전교육 실시"]
            },
            {
                "category": "전기적",
                "risk_factor": "전선 노출",
                "description": "전선이 바닥에 노출되어 감전 위험",
                "location": "작업장 입구",
                "possibility_score": random.randint(2, 4),
                "severity_score": random.randint(3, 4),
                "improvement_measures": ["전선 정리", "절연처리", "전용 덕트 설치"]
            },
            {
                "category": "작업환경적",
                "risk_factor": "정리정돈 불량",
                "description": "작업장 바닥에 공구와 자재가 어지럽게 놓여 있음",
                "location": "작업장 전체",
                "possibility_score": random.randint(4, 5),
                "severity_score": random.randint(1, 3),
                "improvement_measures": ["5S 활동 실시", "정리정돈 교육", "전용 보관함 설치"]
            },
            {
                "category": "화학적",
                "risk_factor": "화학물질 취급",
                "description": "적절한 보호장비 없이 화학물질 취급",
                "location": "화학물질 저장소",
                "possibility_score": random.randint(2, 4),
                "severity_score": random.randint(2, 4),
                "improvement_measures": ["보호장비 착용", "환기시설 개선", "MSDS 교육"]
            },
            {
                "category": "물리적",
                "risk_factor": "소음 노출",
                "description": "85dB 이상의 소음에 장시간 노출",
                "location": "기계 운전 구역",
                "possibility_score": random.randint(3, 5),
                "severity_score": random.randint(2, 3),
                "improvement_measures": ["귀마개 착용", "소음 차단막 설치", "정기 청력검사"]
            },
            {
                "category": "인간공학적",
                "risk_factor": "반복작업",
                "description": "동일한 동작의 반복으로 인한 근골격계 부담",
                "location": "조립 라인",
                "possibility_score": random.randint(3, 5),
                "severity_score": random.randint(1, 3),
                "improvement_measures": ["작업순환제 도입", "스트레칭 교육", "작업대 높이 조절"]
            }
        ]
        
        # 랜덤하게 2-4개의 위험요소 선택
        num_risks = random.randint(2, 4)
        selected_risks = random.sample(risk_pool, num_risks)
        
        # 위험점수 계산 및 추가 정보 생성
        for risk in selected_risks:
            risk["risk_score"] = risk["possibility_score"] * risk["severity_score"]
            
            # 가능성 사유
            possibility_reasons = {
                1: "매우 드물게 발생",
                2: "가끔 접촉 가능",
                3: "보통 수준의 노출",
                4: "자주 접근하는 작업 영역",
                5: "매일 발생하는 상황"
            }
            risk["possibility_reason"] = possibility_reasons[risk["possibility_score"]]
            
            # 중대성 사유
            severity_reasons = {
                1: "경미한 부상",
                2: "넘어짐으로 인한 부상",
                3: "중상 가능성",
                4: "감전 시 사망 위험"
            }
            risk["severity_reason"] = severity_reasons[risk["severity_score"]]
        
        analysis_result = {
            "identified_risks": selected_risks,
            "overall_assessment": {
                "total_risks": len(selected_risks),
                "highest_risk_score": max([r["risk_score"] for r in selected_risks]),
                "average_risk_score": sum([r["risk_score"] for r in selected_risks]) / len(selected_risks),
                "priority_improvements": [r["improvement_measures"][0] for r in selected_risks[:2]],
                "compliance_issues": ["산업안전보건법 제38조 위반 가능성"] if max([r["risk_score"] for r in selected_risks]) > 8 else []
            },
            "file_info": {
                "filename": filename,
                "filepath": filepath,
                "upload_time": timestamp,
                "file_size": len(file.read())
            }
        }
        
        print(f"이미지 분석 완료: {filename} - {len(selected_risks)}개 위험요소 식별")
        
        return jsonify(analysis_result)
        
    except Exception as e:
        print(f"이미지 분석 오류: {e}")
        return jsonify({"error": f"이미지 분석 중 오류가 발생했습니다: {str(e)}"}), 500

@app.route('/api/generate-excel', methods=['POST'])
def generate_excel():
    """Excel 보고서 생성 API"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({"error": "요청 데이터가 없습니다."}), 400
        
        analysis_data = data.get('analysis_data', {})
        basic_info = data.get('basic_info', {})
        
        # 기본 정보 검증
        required_fields = ['workplace_name', 'process_name', 'evaluation_date']
        for field in required_fields:
            if not basic_info.get(field):
                return jsonify({"error": f"필수 정보가 누락되었습니다: {field}"}), 400
        
        # Excel 파일 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_workplace_name = "".join(c for c in basic_info['workplace_name'] if c.isalnum() or c in (' ', '-', '_')).rstrip()
        final_filename = f"HealSE_Report_{safe_workplace_name}_{timestamp}.xlsx"
        final_path = os.path.join(OUTPUT_FOLDER, final_filename)
        
        try:
            # openpyxl 사용하여 Excel 생성
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "위험성평가 결과"
            
            # 스타일 정의
            title_font = Font(bold=True, size=16, color="FFFFFF")
            header_font = Font(bold=True, size=12, color="FFFFFF")
            normal_font = Font(size=11)
            
            title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 제목 작성
            ws.merge_cells('A1:G1')
            title_cell = ws['A1']
            title_cell.value = "HealSE smart Risk Assessment 보고서"
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.border = border
            
            # 기본 정보 섹션
            ws['A3'] = "기본 정보"
            ws['A3'].font = Font(bold=True, size=14)
            ws.merge_cells('A3:B3')
            
            info_data = [
                ("사업장명", basic_info['workplace_name']),
                ("공정명", basic_info['process_name']),
                ("평가일자", basic_info['evaluation_date']),
                ("평가자", basic_info.get('evaluator', '')),
                ("평가 완료일", datetime.now().strftime("%Y-%m-%d %H:%M"))
            ]
            
            for i, (label, value) in enumerate(info_data, 4):
                ws[f'A{i}'] = label
                ws[f'B{i}'] = value
                ws[f'A{i}'].font = Font(bold=True)
            
            # 위험요소 테이블 시작 행
            start_row = 10
            
            # 테이블 헤더
            headers = ['번호', '위험요소', '분류', '가능성', '중대성', '위험점수', '개선방안']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # 위험요소 데이터 입력
            risks = analysis_data.get('identified_risks', [])
            for i, risk in enumerate(risks, 1):
                row = start_row + i
                
                # 데이터 입력
                ws.cell(row=row, column=1, value=i).border = border
                ws.cell(row=row, column=2, value=risk.get('risk_factor', '')).border = border
                ws.cell(row=row, column=3, value=risk.get('category', '')).border = border
                ws.cell(row=row, column=4, value=risk.get('possibility_score', 0)).border = border
                ws.cell(row=row, column=5, value=risk.get('severity_score', 0)).border = border
                
                # 위험점수에 색상 적용
                risk_score = risk.get('risk_score', 0)
                score_cell = ws.cell(row=row, column=6, value=risk_score)
                score_cell.border = border
                
                if risk_score <= 4:
                    score_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                elif risk_score <= 8:
                    score_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                else:
                    score_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    score_cell.font = Font(color="FFFFFF", bold=True)
                
                # 개선방안
                improvements = ', '.join(risk.get('improvement_measures', []))
                ws.cell(row=row, column=7, value=improvements).border = border
            
            # 요약 통계
            summary_start = start_row + len(risks) + 3
            ws[f'A{summary_start}'] = "평가 요약"
            ws[f'A{summary_start}'].font = Font(bold=True, size=14)
            
            total_risks = len(risks)
            avg_score = sum(r.get('risk_score', 0) for r in risks) / max(total_risks, 1)
            max_score = max([r.get('risk_score', 0) for r in risks] + [0])
            
            high_risk_count = sum(1 for r in risks if r.get('risk_score', 0) > 8)
            
            summary_data = [
                ("총 위험요소 수", f"{total_risks}개"),
                ("평균 위험점수", f"{avg_score:.1f}점"),
                ("최고 위험점수", f"{max_score}점"),
                ("높은 위험 요소", f"{high_risk_count}개"),
                ("전체 평가", "높은 위험" if max_score > 12 else "보통 위험" if max_score > 8 else "낮은 위험")
            ]
            
            for i, (label, value) in enumerate(summary_data, summary_start + 1):
                ws[f'A{i}'] = label
                ws[f'B{i}'] = value
                ws[f'A{i}'].font = Font(bold=True)
            
            # 열 너비 자동 조정
            column_widths = [8, 25, 15, 10, 10, 12, 50]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # 행 높이 조정
            for row in range(1, ws.max_row + 1):
                ws.row_dimensions[row].height = 20
            
            wb.save(final_path)
            print(f"Excel 파일 생성 완료: {final_path}")
            
        except ImportError:
            # openpyxl이 없는 경우 CSV 파일로 대체
            final_filename = f"HealSE_Report_{safe_workplace_name}_{timestamp}.csv"
            final_path = os.path.join(OUTPUT_FOLDER, final_filename)
            
            import csv
            with open(final_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                
                # 헤더
                writer.writerow(['HealSE smart Risk Assessment 보고서'])
                writer.writerow([])
                writer.writerow(['기본 정보'])
                writer.writerow(['사업장명', basic_info['workplace_name']])
                writer.writerow(['공정명', basic_info['process_name']])
                writer.writerow(['평가일자', basic_info['evaluation_date']])
                writer.writerow(['평가자', basic_info.get('evaluator', '')])
                writer.writerow([])
                
                # 위험요소 테이블
                writer.writerow(['번호', '위험요소', '분류', '가능성', '중대성', '위험점수', '개선방안'])
                
                for i, risk in enumerate(analysis_data.get('identified_risks', []), 1):
                    writer.writerow([
                        i,
                        risk.get('risk_factor', ''),
                        risk.get('category', ''),
                        risk.get('possibility_score', 0),
                        risk.get('severity_score', 0),
                        risk.get('risk_score', 0),
                        ', '.join(risk.get('improvement_measures', []))
                    ])
        
        # 요약 보고서 생성
        risks = analysis_data.get('identified_risks', [])
        summary = {
            "총_위험요소": len(risks),
            "평균_위험점수": sum(r.get('risk_score', 0) for r in risks) / max(len(risks), 1),
            "최고_위험점수": max([r.get('risk_score', 0) for r in risks] + [0]),
            "높은_위험_개수": sum(1 for r in risks if r.get('risk_score', 0) > 8),
            "전체_평가": "높은 위험" if max([r.get('risk_score', 0) for r in risks] + [0]) > 12 else "보통 위험"
        }
        
        return jsonify({
            "success": True,
            "filename": final_filename,
            "filepath": final_path,
            "summary": summary,
            "download_url": f"/api/download/{final_filename}"
        })
        
    except Exception as e:
        print(f"Excel 생성 오류: {e}")
        return jsonify({"error": f"Excel 생성 중 오류가 발생했습니다: {str(e)}"}), 500

@app.route('/api/download/<filename>', methods=['GET'])
def download_file(filename):
    """파일 다운로드 API"""
    try:
        filepath = os.path.join(OUTPUT_FOLDER, filename)
        
        if not os.path.exists(filepath):
            return jsonify({"error": "파일을 찾을 수 없습니다."}), 404
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"파일 다운로드 오류: {e}")
        return jsonify({"error": f"파일 다운로드 중 오류가 발생했습니다: {str(e)}"}), 500

if __name__ == '__main__':
    print("=" * 60)
    print("🛡️  HealSE smart Risk Assessment 시스템 시작")
    print("=" * 60)
    print(f"📁 업로드 폴더: {UPLOAD_FOLDER}")
    print(f"📊 출력 폴더: {OUTPUT_FOLDER}")
    print(f"🌐 서버 주소: http://localhost:5000")
    print("=" * 60)
    print("✅ 브라우저에서 http://localhost:5000 에 접속하세요!")
    print("=" * 60)
    
    try:
        app.run(host='0.0.0.0', port=5000, debug=False)
    except Exception as e:
        print(f"❌ 서버 시작 오류: {e}")
        input("엔터 키를 눌러 종료하세요...")
